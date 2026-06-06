"""
Сервис экспорта результатов ранжирования.
Содержит только бизнес-логику: формирование Excel-файла.
Автор: Лосева Е.А.
Дата создания: 13.03.2026
Последнее изменение: 01.06.2026
Контакт: ekaterinaloseva91@gmail.com
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:
    """
    Экспорт результатов ранжирования в Excel.
    Методы:
        build_excel_results — формирует Excel-файл с результатами и возвращает байты.
    """

    def build_excel_results(self, results: list) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Результаты'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='0F2747', end_color='0F2747', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col, header in enumerate(['Место', 'ID перевозчика', 'Название', 'Оценка'], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, thin

        for row_idx, r in enumerate(results, 2):
            for col, value in enumerate([r.rank, r.carrier_id, r.carrier.company_name, round(r.score, 4)], 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment, cell.border = center, thin

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()